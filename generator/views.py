from django.shortcuts import render
from django.http import HttpResponse
import datetime
import calendar
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from czech_holidays import holidays

import urllib
import json
from django.conf import settings
import requests



# Create your views here.
def help(request):
    return render(request,'generator/help.html')

def index(request):
    days_in_month = calendar.monthrange(datetime.datetime.now().year, datetime.datetime.now().month)
    days_list = range(1, days_in_month[1] + 1)
    first_day_in_m = range(0, datetime.date(datetime.datetime.now().year, datetime.datetime.now().month, 1).weekday())
    if request.method == 'POST':
        # ''' Begin reCAPTCHA validation '''
        # recaptcha_response = request.POST.get('g-recaptcha-response')
        # url = 'https://www.google.com/recaptcha/api/siteverify'
        # values = {
        #     'secret': settings.GOOGLE_RECAPTCHA_SECRET_KEY,
        #     'response': recaptcha_response
        # }
        # data = urllib.parse.urlencode(values).encode()
        # req = urllib.request.Request(url, data=data)
        # response = urllib.request.urlopen(req)
        # result = json.loads(response.read().decode())
        # ''' End reCAPTCHA validation '''

        #použití lib requests
        ''' Begin reCAPTCHA validation '''
        recaptcha_response = request.POST.get('g-recaptcha-response')
        data = {
            'secret': settings.GOOGLE_RECAPTCHA_SECRET_KEY,
            'response': recaptcha_response
        }
        r = requests.post('https://www.google.com/recaptcha/api/siteverify', data=data)
        result = r.json()
        ''' End reCAPTCHA validation '''


        if result['success']:
            # vygeneruje AIS
            vygenerovana_ais = cruncher(dict(request.POST))

            # uloží AIS
            filename = "AIS" + "_" + str(datetime.date(datetime.datetime.now().year, datetime.datetime.now().month, 1))
            response = HttpResponse(save_virtual_workbook(vygenerovana_ais), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % filename

            vygenerovana_ais.save(response)
            return response
        else:
            return render(request, 'generator/index.html', {'days_list': days_list, "first_day_in_m": first_day_in_m})

    return render(request, 'generator/index.html', {'days_list': days_list, "first_day_in_m": first_day_in_m})


def cruncher(data_dict):
    """main function to create AIS"""

    # načte šablonu a aktivní list
    wb = load_workbook('generator/AIS_master.xlsx')
    ws = wb.active

    def month_name_cz(month_num):
        month_names = ["leden", "únor", "březen", "duben", "květen", "červen",
                       "červenec", "srpen", "září", "říjen", "listopad", "prosinec"]
        return month_names[month_num - 1]

    # vyplní hlavičku šablony
    ws[
        "A5"] = "pracovní doba:                                                                                   za měsíc: %s" \
                % (month_name_cz(datetime.datetime.now().month) + " " + str(datetime.datetime.now().year))

    # vstupní data
    data = data_dict['date']
    data_clone = data

    # iteruje vstupními daty, modifikuje hodnotu dne po noční službě na "z_or_n1_next"
    for day_num, day in enumerate(data_clone):
        if day in ["z", "n1"] and day_num+1 != len(data_clone):
            data_clone[day_num + 1] = "z_or_n1_next"
    data = data_clone

    day_obj_list = []

    class Day():
        """objekt reprezentující každý den v měsící"""

        def __init__(self, day_num, occupation):
            self.day_num = day_num + 1  # number of the day in the month.
            self.occupation = occupation  # ev. "služba" at given day -d1,d2,d3,z,n1

            # today = datetime.date()
            self.date = datetime.date(datetime.datetime.now().year, datetime.datetime.now().month, self.day_num)
            self.den_v_tydnu = self.date.weekday()
            self.svatek = self.date in holidays
            self.vikend = self.den_v_tydnu >= 5
            # self.post_night_shift = False

            self.input = []
            self.to_iter = []

        def populator(self, ais_workbook_ws):
            """funkce, která podle charakteristik objektu vyplňuje šablonu AIS"""

            if self.occupation in ["d1", "d2"]:
                if self.vikend and not self.svatek:
                    self.input = [" 07:00", "19:00", 12, 12, 12]
                    self.to_iter = ["C", "D", "J", "R", "T"]
                elif self.svatek and not self.vikend:
                    self.input = [" 07:00", "19:00", 12, 12, 12]
                    self.to_iter = ["C", "D", "J", "O", "U"]
                elif self.vikend and self.svatek:
                    self.input = [" 07:00", "19:00", 12, 12, 12, 12]
                    self.to_iter = ["C", "D", "J", "R", "T", "U"]
                else:
                    self.input = [" 07:00", "19:00", "11:30", "12:00", 11.5, 3.5]
                    self.to_iter = ["C", "D", "E", "F", "J", "O"]
            elif self.occupation == "d3":
                if self.svatek and self.vikend:
                    self.input = [" 07:00", "19:00", 4, 4, 4, 4]
                    self.to_iter = ["C", "D", "J", "R", "T", "U"]
                elif self.svatek and not self.vikend:
                    self.input = [" 07:00", "19:00", 4, 4, 4]
                    self.to_iter = ["C", "D", "J", "O", "U"]
                else:
                    self.input = [" 07:00", "11:00", 4, 4, 4]
                    self.to_iter = ["C", "D", "J", "R", "T"]

            elif self.occupation in ["z", "n1"]:
                if self.vikend and not self.svatek:
                    self.input = [" 19:00", "00:00", 5, 5, 2, 5]
                    self.to_iter = ["C", "D", "J", "R", "S", "T"]
                elif self.svatek and not self.vikend:
                    self.input = [" 19:00", "00:00", 5, 5, 2, 5]
                    self.to_iter = ["C", "D", "J", "O", "S", "U"]
                elif self.vikend and self.svatek:
                    self.input = [" 19:00", "00:00", 5, 5, 2, 5, 5]
                    self.to_iter = ["C", "D", "J", "R", "S", "T", "U"]
                else:
                    self.input = [" 07:00", "00:00", "11:00", "19:00", 9, 2]
                    self.to_iter = ["C", "D", "E", "F", "J", "S"]
            elif self.occupation == "z_or_n1_next":
                if self.vikend and not self.svatek:
                    self.input = [" 00:00", "07:00", 7, 7, 6, 7]
                    self.to_iter = ["C", "D", "J", "R", "S", "T"]
                elif self.svatek and not self.vikend:
                    self.input = [" 00:00", "07:00", 7, 7, 6, 7]
                    self.to_iter = ["C", "D", "J", "O", "S", "U"]
                elif self.vikend and self.svatek:
                    self.input = [" 00:00", "07:00", 7, 7, 6, 7, 7]
                    self.to_iter = ["C", "D", "J", "R", "S", "T", "U"]
                else:
                    self.input = [" 00:00", "07:00", 7, 6]
                    self.to_iter = ["C", "D", "J", "S"]
            elif self.occupation == "nan":
                if self.vikend and not self.svatek:
                    self.input = []
                    self.to_iter = []
                elif self.svatek and not self.vikend:
                    self.input = [" 07:00", "15:30", "11:30", "12:00", 8]
                    self.to_iter = ["C", "D", "E", "F", "J"]
                elif self.vikend and self.svatek:
                    self.input = []
                    self.to_iter = []
                else:
                    self.input = [" 07:00", "15:30", "11:30", "12:00", 8]
                    self.to_iter = ["C", "D", "E", "F", "J"]

            # Proveden změny v šabloně pro daný den.
            for column_num in range(len(self.to_iter)):
                sel_cell = self.to_iter[column_num] + str(20 + self.day_num)
                ais_workbook_ws[sel_cell] = self.input[column_num]

    for data_index in range(len(data)):
        day_obj_list.append(Day(data_index, data[data_index]))

    for day_obj in day_obj_list:
        day_obj.populator(ws)

    # wb.save("sample.xlsx")

    # uloží file na disk uživatele
    # response = HttpResponse(excel_file.read(),
    #                         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #
    # # set the file name in the Content-Disposition header
    # response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % normal_filename

    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'

    # response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')

    # wb.save(response)
    return wb

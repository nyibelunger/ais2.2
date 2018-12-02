from openpyxl import load_workbook
import datetime
from czech_holidays import holidays
#wb = load_workbook('AIS_master.xlsx')

# grab the active worksheet
#ws = wb.active

# # Data can be assigned directly to cells
# ws['B21'] = 'd2'
# ws['C21'] = '07:00'
# ws['D21'] = '15:30'
# ws['E21'] = '11:30'
# ws['F21'] = '12:00'
#
# print(wb.sheetnames)
#
#
# # Save the file
# #wb.save("sample.xlsx")
# print("saved")
class Day():
    """objekt reprezentující každý den v měsící"""

    def __init__(self, day_num):
        self.day_num = day_num + 1  # number of the day in the month.
        self.date = datetime.date(datetime.datetime.now().year, datetime.datetime.now().month, self.day_num)
        self.den_v_tydnu = self.date.weekday()
        self.svatek = self.date in holidays
        self.vikend = self.den_v_tydnu >= 5
        #self.post_night_shift = False



xyz = Day(1)

print(xyz.vikend)